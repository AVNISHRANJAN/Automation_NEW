"""
reporting/excel_reporter.py — Enterprise-grade Excel test execution report.

Generates a professional .xlsx report from ActionRecord data captured during a
crawl run. Reuses the existing MetadataLogger records — zero new data collection.

Design:
  - One class, ExcelReporter, mirrors the interface of ReportBuilder exactly.
  - build(records, visited_pages) → str (path)
  - All openpyxl object construction is DEFERRED to build() — never at import time.
    This means the module can be imported safely even when openpyxl is not yet
    installed (the try/except sets _OPENPYXL_AVAILABLE=False and build() returns "").
  - StepGenerator is pure (no I/O) — all action→text translation logic lives here.

Column layout:
  A  Test Case ID      | B  Test Steps       | C  Input Data
  D  Expected Results  | E  Actual Results   | F  Test Environment
  G  Execution Status  | H  Bug Severity     | I  Bug Priority
  J  Notes             | K  Screenshot
"""

import logging
import os
import sys
from pathlib import Path
from datetime import datetime, timezone
from typing import List
from urllib.parse import urlparse

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

import config

# ── Optional dependency — never crash at import time ──────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


# ── Column definitions — pure Python, safe at import time ────────────────────
# (header text, column width in Excel units)
_COLUMNS = [
    ("Test Case ID",      12),
    ("Test Steps",        40),
    ("Input Data",        22),
    ("Expected Results",  35),
    ("Actual Results",    35),
    ("Test Environment",  28),
    ("Execution Status",  14),
    ("Bug Severity",      13),
    ("Bug Priority",      12),
    ("Notes",             28),
    ("Screenshot",        36),
]
_NUM_COLS = len(_COLUMNS)

# ── Hex colour palette — plain strings, safe at import time ──────────────────
_HEADER_BG      = "1E3A5F"   # dark navy
_HEADER_FG      = "FFFFFF"   # white
_PASS_BG        = "E8F5E9"   # light green
_FAIL_BG        = "FFEBEE"   # light red
_ALT_BG         = "F4F6F9"   # alternating row tint
_WHITE_BG       = "FFFFFF"
_PASS_FG        = "1B5E20"   # dark green text
_FAIL_FG        = "B71C1C"   # dark red text
_BORDER_COLOR   = "BDBDBD"   # medium grey


# ══════════════════════════════════════════════════════════════════════════════
# Style factory — called ONCE inside build(), never at module load time.
# Returns a plain namespace object whose attributes are openpyxl style objects.
# ══════════════════════════════════════════════════════════════════════════════

class _StyleBundle:
    """
    Holds all openpyxl style objects created in one place.
    Instantiated once per build() call so construction cost is paid only when
    a report is actually generated, and never during module import.
    """
    __slots__ = (
        "header_font", "header_fill", "header_align",
        "data_font", "data_font_b", "data_align", "center_align",
        "pass_font", "fail_font",
        "pass_fill", "fail_fill", "alt_fill", "white_fill",
        "thin_border",
    )

    def __init__(self):
        _side = Side(style="thin", color=_BORDER_COLOR)

        self.thin_border  = Border(left=_side, right=_side, top=_side, bottom=_side)

        self.header_font  = Font(name="Calibri", bold=True, color=_HEADER_FG, size=10)
        self.header_fill  = PatternFill("solid", fgColor=_HEADER_BG)
        self.header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        self.data_font    = Font(name="Calibri", size=9)
        self.data_font_b  = Font(name="Calibri", size=9, bold=True)
        self.data_align   = Alignment(horizontal="left", vertical="top", wrap_text=True)
        self.center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        self.pass_font    = Font(name="Calibri", size=9, bold=True, color=_PASS_FG)
        self.fail_font    = Font(name="Calibri", size=9, bold=True, color=_FAIL_FG)

        self.pass_fill    = PatternFill("solid", fgColor=_PASS_BG)
        self.fail_fill    = PatternFill("solid", fgColor=_FAIL_BG)
        self.alt_fill     = PatternFill("solid", fgColor=_ALT_BG)
        self.white_fill   = PatternFill("solid", fgColor=_WHITE_BG)


# ══════════════════════════════════════════════════════════════════════════════
# Step / Expected / Actual text generation
# ══════════════════════════════════════════════════════════════════════════════

class StepGenerator:
    """
    Converts raw ActionRecord fields into human-readable test step text.
    All methods are pure — no I/O, no side effects, easy to unit test.
    """

    # action key → (step template, expected template)
    _ACTION_MAP = {
        "click":               ("Click {label}",               "{label} should respond to click"),
        "check":               ("Check {label} checkbox",      "Checkbox should be checked"),
        "check_then_uncheck":  ("Toggle {label} checkbox",     "Checkbox should check then uncheck correctly"),
        "radio_select":        ("Select {label} radio option", "Radio option should be selected exclusively"),
        "select":              ("Select option from {label}",  "Dropdown should accept the selection"),
        "fill":                ("Fill {label} field",          "Input field should accept the value"),
        "tab_click":           ("Click {label} tab",           "Tab panel should activate"),
        "link_recorded":       ("Verify {label} link",         "Link should be reachable"),
        "file_upload":         ("Upload file to {label}",      "File should upload successfully"),
        "navigate":            ("Navigate to page",            "Page should load without errors"),
        "skipped_password":    ("Skip {label} (password)",     "Password field is not auto-filled"),
        "skipped_destructive": ("Skip {label} (destructive)",  "Destructive action safely skipped"),
        "skipped_stale":       ("Skip {label} (stale)",        "Element handle refreshed on next cycle"),
        "failed":              ("Interact with {label}",       "Element should respond to interaction"),
    }

    @classmethod
    def step(cls, label: str, action: str, element_type: str) -> str:
        key, _extra = cls._parse_action(action)
        label       = label or element_type or "element"
        template    = cls._ACTION_MAP.get(key, ("Interact with {label}", ""))[0]
        return template.format(label=label)

    @classmethod
    def expected(cls, label: str, action: str, element_type: str) -> str:
        key, _extra = cls._parse_action(action)
        label       = label or element_type or "element"
        template    = cls._ACTION_MAP.get(key, ("", "Interaction should succeed"))[1]
        return template.format(label=label)

    @classmethod
    def actual(cls, record) -> str:
        """Build a human-readable actual-result string from the record."""
        action = record.action
        key, extra = cls._parse_action(action)

        if not record.success:
            msg = record.error_message or record.error_type or "Unknown error"
            return f"FAILED — {msg[:120]}"

        if key == "fill" and extra:
            return f'"{extra}" entered successfully'
        if key == "select" and extra:
            return f'Option "{extra}" selected'
        if key in ("check", "check_then_uncheck"):
            return "Checkbox toggled successfully"
        if key == "radio_select":
            verdict = extra.split(":")[0] if ":" in extra else extra
            return f"Radio option selected — {verdict}"
        if key == "click":
            if "dialog" in extra:
                return f"Click triggered {extra} dialog"
            if "tab" in extra:
                return "New tab opened and closed cleanly"
            return "Click performed successfully"
        if key == "file_upload":
            if "manual_selected" in extra:
                inner = extra.replace("manual_selected(", "").rstrip(")")
                return f"File uploaded: {inner}"
            return "File upload initiated"
        if key == "tab_click":
            return f'Tab "{extra or record.element_label}" activated'
        if key == "link_recorded":
            return "Link recorded for crawl"
        if key == "navigate":
            return "Page loaded successfully"
        if key.startswith("skipped"):
            return f"Skipped: {key.replace('_', ' ')}"

        return action[:100] if action else "Action performed"

    @classmethod
    def input_data(cls, record) -> str:
        """Extract the input value from the action string, if any."""
        key, extra = cls._parse_action(record.action)
        if key == "fill" and extra:
            return extra[:80]
        if key == "select" and extra:
            return extra[:80]
        if key == "file_upload" and "manual_selected" in extra:
            return extra.replace("manual_selected(", "").rstrip(")")[:80]
        return ""

    @staticmethod
    def _parse_action(action: str):
        """
        Split "fill:test@example.com" → ("fill", "test@example.com").
        Handles multi-word prefixes like "check_then_uncheck", "radio_select".
        """
        key, extra = (action.split(":", 1) if ":" in action else (action, ""))
        # Map compound keys to their canonical entry
        for prefix in (
            "skipped_stale", "skipped_password", "skipped_destructive",
            "check_then_uncheck", "radio_select", "tab_click",
            "file_upload", "link_recorded",
        ):
            if key.startswith(prefix):
                return prefix, extra
        return key.lower(), extra


# ══════════════════════════════════════════════════════════════════════════════
# Severity / Priority helpers
# ══════════════════════════════════════════════════════════════════════════════

def _severity(record) -> str:
    if record.success:
        return ""
    et = (record.error_type or "").lower()
    if any(k in et for k in ("navigation_exception", "http_5", "broken_page")):
        return "Critical"
    if any(k in et for k in ("http_4", "interaction_error")):
        return "Major"
    return "Minor"


def _priority(record) -> str:
    return {"Critical": "P1", "Major": "P2", "Minor": "P3"}.get(_severity(record), "")


# ══════════════════════════════════════════════════════════════════════════════
# Main reporter class
# ══════════════════════════════════════════════════════════════════════════════

class ExcelReporter:
    """
    Builds a professional .xlsx test execution report.

    Usage (mirrors ReportBuilder):
        reporter = ExcelReporter(run_id, target_url)
        excel_path = reporter.build(all_records, visited_pages)
    """

    def __init__(self, run_id: str, target_url: str):
        self.run_id     = run_id
        self.target_url = target_url
        report_dir = config.OUTPUT_DIR / "reports"
        report_dir.mkdir(parents=True, exist_ok=True)
        self.report_path = report_dir / f"{run_id}_execution_report.xlsx"

    def build(self, records: list, visited_pages: list) -> str:
        """
        Generate the Excel report and write it to disk.
        Returns the file path as a string, or empty string on failure.
        """
        if not _OPENPYXL_AVAILABLE:
            logger.error(
                "openpyxl not installed — Excel report skipped. "
                "Run:  pip install openpyxl"
            )
            return ""

        try:
            # Build style bundle here — deferred, never at import time
            styles = _StyleBundle()

            wb  = Workbook()
            ws  = wb.active
            ws.title = "Test Execution Report"

            self._write_headers(ws, styles)
            self._write_rows(ws, records, styles)
            self._apply_sheet_settings(ws)

            wb.save(str(self.report_path))
            logger.info("Excel report written: %s", self.report_path)
            return str(self.report_path)

        except Exception as exc:
            logger.error("Failed to write Excel report: %s", exc)
            return ""

    # ── Header row ────────────────────────────────────────────────────────────

    def _write_headers(self, ws, s: _StyleBundle) -> None:
        for col_idx, (header, width) in enumerate(_COLUMNS, start=1):
            cell            = ws.cell(row=1, column=col_idx, value=header)
            cell.font       = s.header_font
            cell.fill       = s.header_fill
            cell.alignment  = s.header_align
            cell.border     = s.thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 28

    # ── Data rows ─────────────────────────────────────────────────────────────

    def _write_rows(self, ws, records: list, s: _StyleBundle) -> None:
        env = self._environment()
        for idx, record in enumerate(records, start=1):
            self._write_data_row(ws, idx + 1, idx, record, env, s, alt=(idx % 2 == 0))

    def _write_data_row(
        self, ws, row: int, tc_idx: int, record, env: str, s: _StyleBundle, alt: bool
    ) -> None:
        sg = StepGenerator

        tc_id      = f"TC_{tc_idx:04d}"
        step       = sg.step(record.element_label, record.action, record.element_type)
        input_val  = sg.input_data(record)
        expected   = sg.expected(record.element_label, record.action, record.element_type)
        actual     = sg.actual(record)
        status     = "PASS" if record.success else "FAIL"
        severity   = _severity(record)
        priority   = _priority(record)
        notes      = self._notes(record)
        screenshot = Path(record.screenshot_path).name if record.screenshot_path else ""

        values = [
            tc_id, step, input_val, expected, actual,
            env, status, severity, priority, notes, screenshot,
        ]

        row_fill = s.fail_fill if not record.success else (s.alt_fill if alt else s.white_fill)

        for col_idx, value in enumerate(values, start=1):
            cell           = ws.cell(row=row, column=col_idx, value=value)
            cell.border    = s.thin_border
            cell.fill      = row_fill

            if col_idx == 1:                    # TC ID — centred, bold
                cell.font      = s.data_font_b
                cell.alignment = s.center_align
            elif col_idx == 7:                  # Status — colour-coded, centred
                cell.font      = s.pass_font if record.success else s.fail_font
                cell.alignment = s.center_align
            elif col_idx in (8, 9):             # Severity, Priority — centred
                cell.font      = s.data_font
                cell.alignment = s.center_align
            else:
                cell.font      = s.data_font
                cell.alignment = s.data_align

        ws.row_dimensions[row].height = 36

    # ── Sheet-level settings ──────────────────────────────────────────────────

    def _apply_sheet_settings(self, ws) -> None:
        ws.freeze_panes           = "A2"
        ws.auto_filter.ref        = f"A1:{get_column_letter(_NUM_COLS)}1"
        ws.page_setup.fitToPage   = True
        ws.page_setup.fitToWidth  = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.orientation = "landscape"
        ws.print_title_rows       = "1:1"

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _environment(self) -> str:
        domain = urlparse(self.target_url).netloc or self.target_url
        return f"Chrome | {domain}"

    @staticmethod
    def _notes(record) -> str:
        parts = []
        if record.timestamp:
            try:
                dt = datetime.strptime(record.timestamp, "%Y-%m-%dT%H:%M:%SZ")
                parts.append(dt.strftime("%H:%M:%S UTC"))
            except Exception:
                parts.append(record.timestamp)
        if record.element_type:
            parts.append(record.element_type)
        if record.error_type and not record.success:
            parts.append(record.error_type)
        return " | ".join(parts)
